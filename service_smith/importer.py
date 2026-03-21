"""Spreadsheet parsing and field mapping."""

from __future__ import annotations

import csv
from pathlib import Path
import re
from typing import Iterable

from service_smith.formats import DEFAULT_ADAPTER


def load_rows(spreadsheet_path: str | Path, field_map: dict[str, str] | None = None) -> list[dict[str, str]]:
    """Load rows from a CSV or spreadsheet-like export into canonical field names."""
    path = Path(spreadsheet_path)
    mapped_fields = field_map or DEFAULT_ADAPTER.field_map

    if path.suffix.lower() == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            return [
                _map_row(row, mapped_fields, row_number=index)
                for index, row in enumerate(reader, start=2)
            ]

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("openpyxl is required to import Excel workbooks.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        rows: list[dict[str, str]] = []
        for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw = {headers[idx]: _stringify(value) for idx, value in enumerate(values) if idx < len(headers)}
            rows.append(_map_row(raw, mapped_fields, row_number=index))
        return rows

    raise ValueError(f"Unsupported spreadsheet format: {path.suffix}")


def _map_row(row: dict[str, str], field_map: dict[str, str], row_number: int) -> dict[str, str]:
    mapped = {
        canonical_name: _stringify(row.get(source_name))
        for canonical_name, source_name in field_map.items()
    }
    mapped["source_row_number"] = str(row_number)
    return _normalize_row(mapped)


def _stringify(value: object) -> str:
    return "" if value is None else str(value).strip()


def preview_rows(rows: Iterable[dict[str, str]], limit: int = 5) -> list[dict[str, str]]:
    return list(rows)[:limit]


def select_rows(
    rows: list[dict[str, str]],
    *,
    row_start: int | None = None,
    row_end: int | None = None,
    limit: int | None = None,
) -> list[dict[str, str]]:
    selected = rows
    if row_start is not None:
        selected = [row for row in selected if _safe_int(row.get("source_row_number")) >= row_start]
    if row_end is not None:
        selected = [row for row in selected if _safe_int(row.get("source_row_number")) <= row_end]
    if limit is not None and limit >= 0:
        selected = selected[:limit]
    return selected


def validate_rows(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    seen_external_ids: set[str] = set()
    seen_keys: set[tuple[str, str, str]] = set()

    for row in rows:
        row_number = row.get("source_row_number", "?")
        if not row.get("customer_name"):
            issues.append({"row": row_number, "level": "error", "message": "Missing customer_name"})
        if not row.get("description") and not row.get("subject"):
            issues.append({"row": row_number, "level": "error", "message": "Missing subject/description"})
        if not row.get("address"):
            issues.append({"row": row_number, "level": "warning", "message": "Missing address"})
        if not row.get("city") or not row.get("state"):
            issues.append({"row": row_number, "level": "warning", "message": "Missing city/state"})
        email = row.get("customer_email", "")
        if email and not _looks_like_email(email):
            issues.append({"row": row_number, "level": "warning", "message": f"Questionable email format: {email}"})
        phone = row.get("customer_phone", "")
        if phone and len(_phone_digits(phone)) < 10:
            issues.append({"row": row_number, "level": "warning", "message": f"Questionable phone format: {phone}"})
        state = row.get("state", "")
        if state and len(state) != 2:
            issues.append({"row": row_number, "level": "warning", "message": f"State should usually be 2 letters: {state}"})
        zip_code = row.get("zip", "")
        if zip_code and not re.fullmatch(r"\d{5}(?:-\d{4})?", zip_code):
            issues.append({"row": row_number, "level": "warning", "message": f"Questionable zip format: {zip_code}"})

        external_id = row.get("external_id", "")
        if external_id:
            if external_id in seen_external_ids:
                issues.append({"row": row_number, "level": "warning", "message": f"Duplicate external_id {external_id}"})
            seen_external_ids.add(external_id)

        key = (
            row.get("customer_name", "").casefold(),
            row.get("address", "").casefold(),
            (row.get("description") or row.get("subject") or "").casefold(),
        )
        if any(key) and key in seen_keys:
            issues.append({"row": row_number, "level": "warning", "message": "Potential duplicate row in import batch"})
        seen_keys.add(key)

    return issues


def _normalize_row(row: dict[str, str]) -> dict[str, str]:
    normalized = dict(row)
    if normalized.get("customer_email"):
        normalized["customer_email"] = normalized["customer_email"].strip().lower()
    if normalized.get("state"):
        normalized["state"] = normalized["state"].strip().upper()
    if normalized.get("zip"):
        normalized["zip"] = normalized["zip"].strip()
    if normalized.get("customer_phone"):
        digits = _phone_digits(normalized["customer_phone"])
        if len(digits) == 10:
            normalized["customer_phone"] = f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}"
        elif digits:
            normalized["customer_phone"] = digits
    return normalized


def _phone_digits(value: str) -> str:
    return "".join(ch for ch in str(value) if ch.isdigit())


def _looks_like_email(value: str) -> bool:
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value.strip()))


def _safe_int(value: str | None) -> int:
    try:
        return int(value or "0")
    except Exception:
        return 0
